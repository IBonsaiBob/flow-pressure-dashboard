#!/usr/bin/env python3
"""
generate_dashboard.py

Generates Flow_Pressure_Dashboard.xlsx

Dashboard layout
────────────────
  Selector area  (rows 1-22, cols A-L)
  ─────────────────────────────────────────────────────────────────────────
    Row 1  : Title
    Row 2  : Column sub-headers
    Rows 3-22: 20 selector rows — one per flow/pressure slot
      Col A  "Flow N ▼" label  |  Col B  flow dropdown  |  Col C  scaling factor
      Col D  flow Δt  |  Col E  flow 💾 save button
      Col F  "Pres N ▼" label  |  Col G  pressure dropdown  |  Col H  offset
      Col I  pres Δt  |  Col J  pres 💾 save button
    Chart controls (cols K-L, rows 3-7): Start/End date filters + Save Rest button
    Row 23 : Notes
    Row 24 : DATA TABLE banner
    Row 25 : Data table column headers
    Row 26+: Formula rows

  Data table (rows 25+, cols A-AO — same sheet, different rows)
  ─────────────────────────────────────────────────────────────────────────
    A     = Date
    B-U   = Flow 1-20 Adjusted   (name from B3-B22 × scale from C3-C22 + Δt from D3-D22)
    V-AO  = Pres 1-20 Adjusted   (name from G3-G22 + offset from H3-H22 + Δt from I3-I22)

  Chart: anchored at M1, floats right — does not overlap selector area.

Usage:
    python3 generate_dashboard.py
"""

import datetime
import io
import os
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule

# ── Colours ────────────────────────────────────────────────────────────────────
DARK_BLUE    = "1F4E79"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "BDD7EE"
DARK_ORANGE  = "C55A11"
LIGHT_ORANGE = "FCE4D6"
WHITE        = "FFFFFF"
LIGHT_GRAY   = "F2F2F2"
DARK_GRAY    = "595959"
GREEN_DARK   = "375623"
GREEN_MID    = "70AD47"
LIGHT_GREEN  = "E2EFDA"
YELLOW_BG    = "FFF2CC"
PURPLE       = "7030A0"

# 20-colour palettes for the up-to-20-flow / up-to-20-pressure chart series
FLOW_COLORS = [
    "2E75B6", "1F4E79", "00B0F0", "0070C0", "5BA3D9",
    "375623", "70AD47", "4EA72C", "595959", "808080",
    "0D47A1", "1565C0", "42A5F5", "039BE5", "00ACC1",
    "00695C", "2E7D32", "558B2F", "37474F", "546E7A",
]
PRES_COLORS = [
    "C55A11", "E67E22", "F39C12", "7030A0", "9B59B6",
    "E74C3C", "922B21", "00695C", "26A69A", "FF6B35",
    "D4380D", "FA8C16", "FFC53D", "52C41A", "13C2C2",
    "1890FF", "722ED1", "EB2F96", "A0522D", "2F4F4F",
]

# ── Sample data (taken directly from the problem statement) ────────────────────
FLOW_NAMES = [
    "AL012", "AL013", "AL014", "AL023", "AL028", "AL029",
    "AL035", "AL036", "AL038", "AL063", "AM005", "AM014",
    "AM015", "AM022", "AM024", "AM026",
]

SAMPLE_DATES = [
    datetime.datetime(2026, 12, 1) + datetime.timedelta(minutes=15 * i)
    for i in range(28)
]

FLOW_ROWS = [
    [3.168205, 2.204250, 2.665153, 3.436250, 3.225157, 2.574550, 3.670000, 2.038000, 3.130835, 4.021250, 22.68599, 19.81207, -999, 16.92209, 16.58575, 18.88019],
    [3.190769, 2.225250, 2.681334, 3.411250, 3.236456, 2.583137, 3.683750, 2.049000, 3.145557, 4.048750, 22.45106, 19.72593, -999, 17.10220, 18.43384, 15.47377],
    [3.216410, 2.246250, 2.674793, 3.387500, 3.262896, 2.597148, 3.707500, 2.060000, 3.148535, 4.073750, 22.20831, 19.66328, -999, 16.90642, 18.70008, 16.69538],
    [3.242051, 2.265750, 2.707717, 3.392500, 3.289562, 2.612967, 3.731250, 2.072000, 3.148144, 4.123750, 21.40173, 19.91387, -999, 16.70282, 18.12843, 13.89194],
    [3.265641, 2.286750, 2.735150, 3.433750, 3.293630, 2.629012, 3.757500, 2.084000, 3.140112, 4.150000, 21.87158, 19.90604, -999, 16.92992, 18.02663, 13.51606],
    [3.289231, 2.308500, 2.761070, 3.472500, 3.295212, 2.642345, 3.783750, 2.096000, 3.131763, 4.201250, 21.92640, 19.55365, -999, 17.04738, 19.00549, 14.44010],
    [3.312820, 2.331000, 2.780839, 3.520000, 3.299054, 2.654774, 3.811250, 2.109000, 3.122534, 4.228750, 21.77761, 19.82773, -999, 16.87510, 18.58262, 13.68834],
    [3.336410, 2.353500, 2.799314, 3.547500, 3.304251, 2.667655, 3.836250, 2.122000, 3.124414, 4.278750, 22.25529, 20.39155, -999, 16.96907, 18.61395, 13.22632],
    [3.361026, 2.376000, 2.818888, 3.568750, 3.311031, 2.679632, 3.863750, 2.134000, 3.126025, 4.305000, 21.56618, 19.71027, -999, 17.07087, 18.55913, 15.30149],
    [3.386667, 2.399250, 2.838779, 3.562500, 3.319166, 2.690253, 3.891250, 2.132000, 3.146167, 4.356250, 22.04386, 20.29758, -999, 17.27448, 18.80189, 13.70400],
    [3.408205, 2.421750, 2.859207, 3.532500, 3.327754, 2.700649, 3.920000, 2.126000, 3.177197, 4.382500, 21.87158, 20.21144, -999, 17.17267, 18.28505, 14.43227],
    [3.432820, 2.443500, 2.879171, 3.486250, 3.337245, 2.711044, 3.946250, 2.122000, 3.214893, 4.433750, 21.90290, 19.25608, -999, 16.87510, 17.63509, 15.83399],
    [3.459487, 2.466000, 2.899551, 3.433750, 3.346736, 2.721213, 3.983750, 2.118000, 3.255225, 4.462500, 22.41191, 19.23259, -999, 16.89076, 18.30854, 13.99374],
    [3.483077, 2.488500, 2.920345, 3.401250, 3.356454, 2.731382, 4.020000, 2.114000, 3.294092, 4.516250, 21.76195, 20.37589, -999, 16.99256, 18.44167, 14.43227],
    [3.509743, 2.510250, 2.940358, 3.396250, 3.366623, 2.741778, 4.056250, 2.109000, 3.317798, 4.541250, 21.66015, 19.38920, -999, 16.71848, 17.90134, 18.26155],
    [3.536410, 2.532000, 2.960762, 3.437500, 3.375436, 2.750817, 4.090000, 2.106000, 3.326123, 4.598750, 21.96555, 19.57714, -999, 16.86727, 18.25373, 14.00940],
    [3.557949, 2.553750, 2.980775, 3.505000, 3.382216, 2.759856, 4.126250, 2.101000, 3.323169, 4.630000, 21.89507, 19.89038, -999, 16.89076, 18.12060, 14.13470],
    [3.582564, 2.576250, 3.000666, 3.542500, 3.384476, 2.770930, 4.161250, 2.097000, 3.337646, 4.640000, 21.56618, 20.18012, -999, 17.06304, 18.63744, 13.35944],
    [3.608205, 2.597250, 3.020386, 3.566250, 3.388544, 2.780873, 4.197500, 2.092000, 3.355957, 4.621250, 21.82459, 20.15663, -999, 17.01606, 18.47299, 12.76430],
    [3.630769, 2.619750, 3.036445, 3.565000, 3.394193, 2.791042, 4.231250, 2.088000, 3.357056, 4.585000, 22.43540, 19.61630, -999, 16.89076, 18.33204, 13.22632],
    [3.655385, 2.640000, 3.027903, 3.533750, 3.398035, 2.799856, 4.265000, 2.083000, 3.353101, 4.567500, 21.37041, 20.14096, -999, 17.11786, 18.79405, 14.00157],
    [3.678974, 2.661000, 3.017896, 3.487500, 3.399164, 2.807991, 4.297500, 2.078000, 3.355786, 4.531250, 22.05169, 20.10181, -999, 17.24315, 17.46281, 18.08145],
    [3.700513, 2.677500, 3.007158, 3.435000, 3.406622, 2.844600, 4.328750, 2.073000, 3.358057, 4.512500, 21.56618, 19.55365, -999, 16.91426, 17.43932, 17.95615],
    [3.720000, 2.696250, 2.996370, 3.396250, 3.372951, 2.861775, 4.352500, 2.070000, 3.367798, 4.476250, 21.23728, 19.95302, -999, 17.01606, 17.38450, 19.17776],
    [3.737436, 2.712000, 2.986949, 3.387500, 3.340183, 2.857933, 4.371250, 2.062000, 3.381812, 4.456250, 20.77526, 19.89821, -999, 16.66367, 17.36101, 18.64527],
    [3.744615, 2.723250, 2.944190, 3.408750, 3.320748, 2.852058, 4.382500, 2.053000, 3.390039, 4.417500, 21.59750, 19.35005, -999, 16.69499, 17.40799, 18.93501],
    [3.749743, 2.730750, 2.942969, 3.447500, 3.304025, 2.842115, 4.386250, 2.056000, 3.391602, 4.396250, 21.48004, 19.64762, -999, 16.38959, 17.84652, 20.49335],
    [3.746667, 2.732250, 2.938479, 3.488750, 3.288884, 2.830589, 4.386250, 2.062000, 3.384717, 4.351250, 21.62099, 18.58262, -999, 16.37393, 18.44167, 20.41504],
]

PRES_ROWS = [
    [round(v * 0.97 + 0.3, 6) if v != -999 else -999 for v in row]
    for row in FLOW_ROWS
]

# ── Layout constants ───────────────────────────────────────────────────────────
#
#  Dashboard selector area  (rows 1-22, cols A-L):
#
#    Row 1  : Title banner
#    Row 2  : Column sub-headers
#    Rows 3-22: 20 selector rows  (rows 3-22 serve both flow[0-19] and pres[0-19])
#      Col A  row-number  |  Col B  flow dropdown  |  Col C  scale  |  Col D  flow Δt  |  Col E  flow 💾
#      Col F  row-number  |  Col G  pressure dropdown  |  Col H  offset  |  Col I  pres Δt  |  Col J  pres 💾
#
#  Chart controls  (rows 2-7, cols K-L — right of selector, left of chart):
#      L3  Start Date filter  |  L4  End Date filter
#      L5  MATCH helper – start raw row  |  L6  MATCH helper – end raw row
#      K-L7  Save Rest button
#
#    Row 23 : Note / instructions
#    Row 24 : DATA TABLE section banner
#    Row 25 : Data table column headers  (formula-based — shows sensor code)
#    Row 26+: Formula rows  (DATE | Flow 1-20 Adj | Pres 1-20 Adj)
#
#  Data table  (rows 25+, cols A-AO  — same sheet, different rows):
#    A     = Date
#    B-U   = Flow 1-20 Adjusted  (driven by B3-B22 name + C3-C22 scale + D3-D22 Δt)
#    V-AO  = Pres 1-20 Adjusted  (driven by G3-G22 name + H3-H22 offset + I3-I22 Δt)
#
#  Chart: anchored at M1, floats to the right — does not overlap control area.

MAX_FLOW = 20
MAX_PRES = 20

TITLE_ROW       = 1
SEL_HDR_ROW     = 2
SEL_START_ROW   = 3                                        # first selector row
SEL_END_ROW     = SEL_START_ROW + max(MAX_FLOW, MAX_PRES) - 1  # = 22
NOTE_ROW        = SEL_END_ROW + 1                          # = 23
DATA_BANNER_ROW = NOTE_ROW + 1                             # = 24
DATA_HDR_ROW    = DATA_BANNER_ROW + 1                      # = 25
DATA_START_ROW  = DATA_HDR_ROW + 1                         # = 26
DATA_OFFSET     = DATA_START_ROW - 2                       # = 24  → ROW()-24=2 at row 26
DATA_ROWS       = 2000

# Selector column indices (1-based)
COL_FLOW_LABEL  = 1   # A  row number
COL_FLOW_SEL    = 2   # B  flow name dropdown
COL_FLOW_SCALE  = 3   # C  scaling factor
COL_FLOW_DT     = 4   # D  per-row timestep offset
COL_FLOW_SAVE   = 5   # E  per-row save button
COL_PRES_LABEL  = 6   # F  row number
COL_PRES_SEL    = 7   # G  pressure name dropdown
COL_PRES_OFFSET = 8   # H  offset
COL_PRES_DT     = 9   # I  per-row timestep offset
COL_PRES_SAVE   = 10  # J  per-row save button

# Data table column indices (rows DATA_HDR_ROW+, same sheet)
COL_DATE          = 1   # A
COL_FLOW_ADJ_BASE = 2   # B = Flow 1 Adj … U = Flow 20 Adj  (index = base + n)
COL_PRES_ADJ_BASE = 22  # V = Pres 1 Adj … AO = Pres 20 Adj (index = base + n)

CHART_ANCHOR     = "M1"   # starts immediately right of control area at K-L
CHART_WIDTH_CM   = 30
CHART_HEIGHT_CM  = 20

# ── Control area (cols K-L, right of selector, rows 2-7) ──────────────────────
CTRL_LABEL_COL      = 11                       # K
CTRL_INPUT_COL      = 12                       # L
CTRL_START_DATE_ROW  = SEL_START_ROW           # 3  – Start Date filter
CTRL_END_DATE_ROW    = SEL_START_ROW + 1       # 4  – End Date filter
CTRL_HELP_START_ROW  = SEL_START_ROW + 2       # 5  – MATCH helper: start raw row
CTRL_HELP_END_ROW    = SEL_START_ROW + 3       # 6  – MATCH helper: end raw row
CTRL_SAVE_REST_ROW   = SEL_START_ROW + 4       # 7  – Save Rest button


# ── Style helpers ──────────────────────────────────────────────────────────────

def _thin():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def _medium():
    m = Side(style="medium")
    return Border(left=m, right=m, top=m, bottom=m)


def style_header(cell, text, bg=DARK_BLUE, fg=WHITE, bold=True, sz=10,
                 halign="center"):
    cell.value = text
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border = _thin()


def style_label(cell, text, bold=False, fg="000000", sz=10,
                halign="left", italic=False):
    cell.value = text
    cell.font = Font(bold=bold, color=fg, size=sz, italic=italic)
    cell.alignment = Alignment(horizontal=halign, vertical="center")


def style_input(cell, value, bg=LIGHT_BLUE, fg="000000", bold=True, sz=11,
                num_fmt=None):
    cell.value = value
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _medium()
    if num_fmt:
        cell.number_format = num_fmt


# ── Raw data sheets ────────────────────────────────────────────────────────────

def build_raw_sheet(ws, title, table_name, data_rows, dates):
    ws.title = title
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 21

    # Headers
    style_header(ws.cell(1, 1), "Date", bg=DARK_BLUE)
    for ci, name in enumerate(FLOW_NAMES, start=2):
        style_header(ws.cell(1, ci), name, bg=MID_BLUE)
        ws.column_dimensions[get_column_letter(ci)].width = 13

    # Data rows
    for ri, (dt, row) in enumerate(zip(dates, data_rows), start=2):
        c = ws.cell(ri, 1, value=dt)
        c.number_format = "DD/MM/YYYY HH:MM"
        c.alignment = Alignment(horizontal="center", vertical="center")
        for ci, val in enumerate(row, start=2):
            dc = ws.cell(ri, ci, value=val)
            dc.number_format = "0.000000"
            dc.alignment = Alignment(horizontal="right")

    # Named Excel Table (makes Power Query setup one-click)
    last_col = get_column_letter(len(FLOW_NAMES) + 1)
    last_row = len(data_rows) + 1
    tbl = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)


# ── Dashboard sheet ────────────────────────────────────────────────────────────

def build_dashboard(ws, flow_names):
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18   # flow row # (rows 1-22) / Date (rows 26+)
    ws.column_dimensions["B"].width = 15   # flow dropdown / flow 1 adj
    ws.column_dimensions["C"].width = 10   # scale / flow 2 adj
    ws.column_dimensions["D"].width = 8    # flow Δt / flow 3 adj
    ws.column_dimensions["E"].width = 5    # flow 💾 save / flow 4 adj
    ws.column_dimensions["F"].width = 6    # pres row # / flow 5 adj
    ws.column_dimensions["G"].width = 15   # pres dropdown / flow 6 adj
    ws.column_dimensions["H"].width = 10   # pres offset / flow 7 adj
    ws.column_dimensions["I"].width = 8    # pres Δt / flow 8 adj
    ws.column_dimensions["J"].width = 5    # pres 💾 save / flow 9 adj
    ws.column_dimensions["K"].width = 14   # chart control labels / flow 10 adj
    ws.column_dimensions["L"].width = 13   # chart control inputs / flow 11 adj
    for col in [get_column_letter(i) for i in range(13, COL_PRES_ADJ_BASE + MAX_PRES + 1)]:
        ws.column_dimensions[col].width = 13

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[TITLE_ROW].height = 30
    tc = ws.cell(TITLE_ROW, 1, value="Flow & Pressure Analysis Dashboard")
    tc.font = Font(bold=True, color=WHITE, size=14)
    tc.fill = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A1:{get_column_letter(CTRL_INPUT_COL)}1")

    # ── Row 2: Sub-headers ────────────────────────────────────────────────────
    ws.row_dimensions[SEL_HDR_ROW].height = 22
    for ci, (txt, bg) in enumerate([
        ("#",       MID_BLUE),
        ("Flow  ▼", MID_BLUE),
        ("Scale",   MID_BLUE),
        ("Δt",      MID_BLUE),
        ("💾",      MID_BLUE),
        ("#",       DARK_ORANGE),
        ("Pres  ▼", DARK_ORANGE),
        ("Offset",  DARK_ORANGE),
        ("Δt",      DARK_ORANGE),
        ("💾",      DARK_ORANGE),
    ], start=1):
        c = ws.cell(SEL_HDR_ROW, ci, value=txt)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()
    # "Chart Controls" header spanning K2:L2
    _cl = get_column_letter(CTRL_LABEL_COL)
    _ci = get_column_letter(CTRL_INPUT_COL)
    ws.merge_cells(f"{_cl}{SEL_HDR_ROW}:{_ci}{SEL_HDR_ROW}")
    cc = ws.cell(SEL_HDR_ROW, CTRL_LABEL_COL, value="Chart Controls")
    cc.font = Font(bold=True, color=WHITE, size=9)
    cc.fill = PatternFill(fill_type="solid", fgColor=PURPLE)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.border = _thin()

    # ── Rows SEL_START_ROW … SEL_END_ROW: up to 20 selector rows ─────────────
    dv_flow = DataValidation(
        type="list",
        formula1="'Raw Flow Data'!$B$1:$GR$1",
        allow_blank=True, showDropDown=False, showErrorMessage=False,
    )
    dv_pres = DataValidation(
        type="list",
        formula1="'Raw Pressure Data'!$B$1:$GR$1",
        allow_blank=True, showDropDown=False, showErrorMessage=False,
    )
    ws.add_data_validation(dv_flow)
    ws.add_data_validation(dv_pres)

    for n in range(max(MAX_FLOW, MAX_PRES)):
        r = SEL_START_ROW + n
        ws.row_dimensions[r].height = 24

        # ── Flow side ──────────────────────────────────────────────────────
        if n < MAX_FLOW:
            lc = ws.cell(r, COL_FLOW_LABEL, value=n + 1)
            lc.font = Font(bold=True, color=DARK_BLUE, size=9)
            lc.alignment = Alignment(horizontal="center", vertical="center")

            sel_cell = ws.cell(r, COL_FLOW_SEL)
            style_input(sel_cell,
                        flow_names[n] if n < len(flow_names) else "",
                        bg=FLOW_COLORS[n], fg="FFFFFF", sz=11)
            dv_flow.add(sel_cell)

            sc = ws.cell(r, COL_FLOW_SCALE)
            style_input(sc, 1.0, bg=LIGHT_ORANGE, num_fmt="0.000")

            dtf = ws.cell(r, COL_FLOW_DT)
            style_input(dtf, 0, bg=YELLOW_BG, num_fmt="0")

            fsave = ws.cell(r, COL_FLOW_SAVE, value="💾")
            fsave.fill = PatternFill(fill_type="solid", fgColor=MID_BLUE)
            fsave.font = Font(color=WHITE, size=9)
            fsave.alignment = Alignment(horizontal="center", vertical="center")
            fsave.border = _thin()

        # ── Pressure side ──────────────────────────────────────────────────
        if n < MAX_PRES:
            lp = ws.cell(r, COL_PRES_LABEL, value=n + 1)
            lp.font = Font(bold=True, color=DARK_ORANGE, size=9)
            lp.alignment = Alignment(horizontal="center", vertical="center")

            psel = ws.cell(r, COL_PRES_SEL)
            style_input(psel,
                        flow_names[n] if n < len(flow_names) else "",
                        bg=PRES_COLORS[n], fg="FFFFFF", sz=11)
            dv_pres.add(psel)

            oc = ws.cell(r, COL_PRES_OFFSET)
            style_input(oc, 0.0, bg=LIGHT_GREEN, num_fmt="0.000")

            dtp = ws.cell(r, COL_PRES_DT)
            style_input(dtp, 0, bg=YELLOW_BG, num_fmt="0")

            psave = ws.cell(r, COL_PRES_SAVE, value="💾")
            psave.fill = PatternFill(fill_type="solid", fgColor=DARK_ORANGE)
            psave.font = Font(color=WHITE, size=9)
            psave.alignment = Alignment(horizontal="center", vertical="center")
            psave.border = _thin()

    # ── Control area (cols K-L, rows 3-7) ────────────────────────────────────
    # Date format DD/MM/YY is intentionally compact (user preference).
    _ctrl_rows = [
        (CTRL_START_DATE_ROW, "Start ▶", "", "DD/MM/YY", LIGHT_BLUE),
        (CTRL_END_DATE_ROW,   "End ▶",   "", "DD/MM/YY", LIGHT_BLUE),
    ]
    for (row, label_txt, default_val, num_fmt, inp_bg) in _ctrl_rows:
        lc = ws.cell(row, CTRL_LABEL_COL, value=label_txt)
        lc.font = Font(bold=True, color=PURPLE, size=9)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        style_input(ws.cell(row, CTRL_INPUT_COL), default_val,
                    bg=inp_bg, num_fmt=num_fmt)

    # Helper formulas — derived from Start/End Date; users should not edit these
    _ic_ltr = get_column_letter(CTRL_INPUT_COL)   # "L"
    _help_rows = [
        (CTRL_HELP_START_ROW, "▸ s.row:",
         f"=IF(${_ic_ltr}${CTRL_START_DATE_ROW}=\"\",2,"
         f"IFERROR(MATCH(${_ic_ltr}${CTRL_START_DATE_ROW},"
         f"'Raw Flow Data'!$A$2:$A$50001,1)+1,2))"),
        (CTRL_HELP_END_ROW, "▸ e.row:",
         f"=IF(${_ic_ltr}${CTRL_END_DATE_ROW}=\"\",9999999,"
         f"IFERROR(MATCH(${_ic_ltr}${CTRL_END_DATE_ROW},"
         f"'Raw Flow Data'!$A$2:$A$50001,1)+1,9999999))"),
    ]
    for (row, label_txt, formula) in _help_rows:
        lc = ws.cell(row, CTRL_LABEL_COL, value=label_txt)
        lc.font = Font(italic=True, color=DARK_GRAY, size=8)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        hc = ws.cell(row, CTRL_INPUT_COL, value=formula)
        hc.font = Font(color=DARK_GRAY, size=8)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.number_format = "0"

    # "Save Rest" button — click to save all sensors not yet in MOD sheets
    _cl = get_column_letter(CTRL_LABEL_COL)
    _ci = get_column_letter(CTRL_INPUT_COL)
    ws.merge_cells(f"{_cl}{CTRL_SAVE_REST_ROW}:{_ci}{CTRL_SAVE_REST_ROW}")
    sr_cell = ws.cell(CTRL_SAVE_REST_ROW, CTRL_LABEL_COL, value="💾 Save Rest")
    sr_cell.fill = PatternFill(fill_type="solid", fgColor=PURPLE)
    sr_cell.font = Font(bold=True, color=WHITE, size=9)
    sr_cell.alignment = Alignment(horizontal="center", vertical="center")
    sr_cell.border = _thin()

    # ── Note row ──────────────────────────────────────────────────────────────
    ws.row_dimensions[NOTE_ROW].height = 44
    nc = ws.cell(NOTE_ROW, 1,
                 value=("ℹ  Up to 20 flow and 20 pressure series.  "
                        "Leave a Name cell blank to hide that series.  "
                        "Each row has its own Scale (flow multiplier), Offset (pressure addend) "
                        "and Δt (integer timestep offset — positive shifts the series later, "
                        "negative shifts it earlier).  "
                        "After pasting your own data, right-click a Name cell → Data Validation "
                        "→ update the Source range to match your column headers.  "
                        "Values of -999 are treated as no-data and excluded."))
    nc.font = Font(italic=True, color=DARK_GRAY, size=9)
    nc.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{NOTE_ROW}:{get_column_letter(COL_PRES_ADJ_BASE + MAX_PRES - 1)}{NOTE_ROW}")

    # ── DATA TABLE section banner ──────────────────────────────────────────────
    ws.row_dimensions[DATA_BANNER_ROW].height = 20
    sc = ws.cell(DATA_BANNER_ROW, 1,
                 value="FORMULA TABLE  —  updates automatically when you change selections or adjustments above")
    sc.font = Font(bold=True, color=WHITE, size=9)
    sc.fill = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A{DATA_BANNER_ROW}:{get_column_letter(COL_PRES_ADJ_BASE + MAX_PRES - 1)}{DATA_BANNER_ROW}")

    # ── DATA TABLE column headers (formula-driven: show sensor code or fallback) ─
    ws.row_dimensions[DATA_HDR_ROW].height = 22
    style_header(ws.cell(DATA_HDR_ROW, COL_DATE), "Date", bg=DARK_BLUE)
    for n in range(MAX_FLOW):
        sel_row = SEL_START_ROW + n
        style_header(ws.cell(DATA_HDR_ROW, COL_FLOW_ADJ_BASE + n), "", bg=MID_BLUE)
        ws.cell(DATA_HDR_ROW, COL_FLOW_ADJ_BASE + n).value = (
            f'=IF($B${sel_row}="","",$B${sel_row})'
        )
    for n in range(MAX_PRES):
        sel_row = SEL_START_ROW + n
        _psel_ltr = get_column_letter(COL_PRES_SEL)
        style_header(ws.cell(DATA_HDR_ROW, COL_PRES_ADJ_BASE + n), "", bg=GREEN_DARK)
        ws.cell(DATA_HDR_ROW, COL_PRES_ADJ_BASE + n).value = (
            f'=IF(${_psel_ltr}${sel_row}="","",${_psel_ltr}${sel_row})'
        )

    # ── DATA TABLE formula rows ────────────────────────────────────────────────
    _sr = CTRL_HELP_START_ROW   # L5 – start raw-data row
    _er = CTRL_HELP_END_ROW     # L6 – end raw-data row
    # Per-series Δt: flow uses $D${sel_row}, pres uses $I${sel_row}
    _ic = get_column_letter(CTRL_INPUT_COL)   # "L"
    _fd = get_column_letter(COL_FLOW_DT)      # "D"
    _pd = get_column_letter(COL_PRES_DT)      # "I"
    _fsel = get_column_letter(COL_FLOW_SEL)    # "B"
    _fscl = get_column_letter(COL_FLOW_SCALE)  # "C"
    _psel = get_column_letter(COL_PRES_SEL)    # "G"
    _poff = get_column_letter(COL_PRES_OFFSET) # "H"
    for r in range(DATA_START_ROW, DATA_START_ROW + DATA_ROWS):
        ws.row_dimensions[r].height = 15
        alt = (r % 2 == 0)
        alt_fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY) if alt else None

        # Col A: Date  (base row = $K$5 + ROW() - DATA_START_ROW)
        ac = ws.cell(r, COL_DATE)
        ac.value = (
            f"=IFERROR("
            f"IF(${_ic}${_sr}+ROW()-{DATA_START_ROW}>${_ic}${_er},NA(),"
            f"IF(INDEX('Raw Flow Data'!$A:$A,${_ic}${_sr}+ROW()-{DATA_START_ROW})=\"\",NA(),"
            f"INDEX('Raw Flow Data'!$A:$A,${_ic}${_sr}+ROW()-{DATA_START_ROW}))),NA())"
        )
        ac.number_format = "DD/MM/YYYY HH:MM"
        ac.alignment = Alignment(horizontal="center")
        if alt_fill:
            ac.fill = alt_fill

        # Cols B-U: Flow 1-20 Adjusted  (row offset per-series from col D)
        for n in range(MAX_FLOW):
            sel_row = SEL_START_ROW + n
            col = COL_FLOW_ADJ_BASE + n
            c = ws.cell(r, col)
            c.value = (
                f"=IFERROR("
                f"IF(${_fsel}${sel_row}=\"\",NA(),"
                f"IF(${_ic}${_sr}+ROW()-{DATA_START_ROW}>${_ic}${_er},NA(),"
                f"IF(INDEX('Raw Flow Data'!$A:$ZZ,${_ic}${_sr}+ROW()-{DATA_START_ROW}-${_fd}${sel_row},"
                f"MATCH(${_fsel}${sel_row},'Raw Flow Data'!$1:$1,0))=-999,NA(),"
                f"INDEX('Raw Flow Data'!$A:$ZZ,${_ic}${_sr}+ROW()-{DATA_START_ROW}-${_fd}${sel_row},"
                f"MATCH(${_fsel}${sel_row},'Raw Flow Data'!$1:$1,0))*${_fscl}${sel_row}))),NA())"
            )
            c.number_format = "0.000"
            c.alignment = Alignment(horizontal="right")
            if alt_fill:
                c.fill = alt_fill

        # Cols V-AO: Pressure 1-20 Adjusted  (row offset per-series from col I)
        for n in range(MAX_PRES):
            sel_row = SEL_START_ROW + n
            col = COL_PRES_ADJ_BASE + n
            c = ws.cell(r, col)
            c.value = (
                f"=IFERROR("
                f"IF(${_psel}${sel_row}=\"\",NA(),"
                f"IF(${_ic}${_sr}+ROW()-{DATA_START_ROW}>${_ic}${_er},NA(),"
                f"IF(INDEX('Raw Pressure Data'!$A:$ZZ,${_ic}${_sr}+ROW()-{DATA_START_ROW}-${_pd}${sel_row},"
                f"MATCH(${_psel}${sel_row},'Raw Pressure Data'!$1:$1,0))=-999,NA(),"
                f"INDEX('Raw Pressure Data'!$A:$ZZ,${_ic}${_sr}+ROW()-{DATA_START_ROW}-${_pd}${sel_row},"
                f"MATCH(${_psel}${sel_row},'Raw Pressure Data'!$1:$1,0))+${_poff}${sel_row}))),NA())"
            )
            c.number_format = "0.000"
            c.alignment = Alignment(horizontal="right")
            if alt_fill:
                c.fill = alt_fill

    # ── Conditional formatting: hide #N/A cells in the data table ─────────────
    # Formulas return NA() instead of "" for empty/out-of-range cells so scatter
    # charts treat them as gaps (not x=0).  This CF rule makes those cells appear
    # blank by applying white text + white fill, overriding the alternating rows.
    cf_last_col = get_column_letter(COL_PRES_ADJ_BASE + MAX_PRES - 1)
    ws.conditional_formatting.add(
        f'A{DATA_START_ROW}:{cf_last_col}{DATA_START_ROW + DATA_ROWS - 1}',
        FormulaRule(
            formula=[f'ISNA(A{DATA_START_ROW})'],
            font=Font(color='FFFFFF'),
            fill=PatternFill(fill_type='solid', fgColor='FFFFFF'),
        )
    )

    # ── Copy-down hint ────────────────────────────────────────────────────────
    hint_r = DATA_START_ROW + DATA_ROWS
    ws.row_dimensions[hint_r].height = 24
    last_data = DATA_START_ROW + DATA_ROWS - 1
    last_col  = get_column_letter(COL_PRES_ADJ_BASE + MAX_PRES - 1)
    hc = ws.cell(hint_r, 1,
                 value=(f"↑ Formulas cover {DATA_ROWS} rows (row {DATA_START_ROW}–{last_data}). "
                        f"For more data: select A{DATA_START_ROW}:{last_col}{last_data} and copy-paste downward."))
    hc.font = Font(italic=True, color=DARK_GRAY, size=9)
    hc.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{hint_r}:{last_col}{hint_r}")

    # ── Chart ──────────────────────────────────────────────────────────────────
    _add_chart(ws)
def _add_chart(ws):
    """Create a minimal placeholder chart that openpyxl embeds in the xlsx.
    The entire chart1.xml is replaced by _patch_chart_xml; this just ensures
    the required xl/charts/ and xl/drawings/ ZIP entries exist."""
    c1 = LineChart()
    c1.title = "Flow & Pressure Analysis"
    c1.style = 10
    c1.width  = CHART_WIDTH_CM
    c1.height = CHART_HEIGHT_CM
    last_row = DATA_START_ROW + DATA_ROWS - 1
    ref = Reference(ws, min_col=COL_FLOW_ADJ_BASE, max_col=COL_FLOW_ADJ_BASE,
                    min_row=DATA_HDR_ROW, max_row=last_row)
    c1.add_data(ref, titles_from_data=True)
    c1.anchor = CHART_ANCHOR
    ws.add_chart(c1)


def _build_correct_chart_xml():
    """Return a valid dual-axis scatter-with-smooth-lines chart XML.

    Scatter (XY) charts require:
      • <c:scatterChart> instead of <c:lineChart>
      • Per-series <c:xVal> (date column) and <c:yVal> (data column)
        instead of <c:cat> and <c:val>
      • X axis defined as <c:valAx> (numeric) rather than <c:catAx>
      • <c:smooth val="1"/> per series for smooth lines

    Data table layout (Dashboard sheet):
      Col A            = Date  (rows DATA_START_ROW – DATA_START_ROW+DATA_ROWS-1)
      Cols B-U (2-21)  = Flow 1-20 Adjusted
      Cols V-AO (22-41)= Pres 1-20 Adjusted

    The chart title of each series is driven by the header-row formula in
    DATA_HDR_ROW, which returns "" when the selector is blank, causing Excel
    to show no legend entry text for that series.
    """
    last_row  = DATA_START_ROW + DATA_ROWS - 1
    date_col  = get_column_letter(COL_DATE)   # A

    def _ref(col_ltr):
        return f"Dashboard!${col_ltr}${DATA_START_ROW}:${col_ltr}${last_row}"

    def _hdr(col_ltr):
        return f"Dashboard!${col_ltr}${DATA_HDR_ROW}"

    empty_cache = (
        '<c:numCache>'
        '<c:formatCode>General</c:formatCode>'
        '<c:ptCount val="0"/>'
        '</c:numCache>'
    )

    def _ser(idx, color, col_ltr, dashed=False):
        dash_xml = '<a:prstDash val="dash"/>' if dashed else ""
        return (
            f'        <c:ser>\n'
            f'          <c:idx val="{idx}"/>\n'
            f'          <c:order val="{idx}"/>\n'
            f'          <c:tx><c:strRef><c:f>{_hdr(col_ltr)}</c:f></c:strRef></c:tx>\n'
            f'          <c:spPr>'
            f'<a:ln w="20000">'
            f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>'
            f'{dash_xml}'
            f'</a:ln></c:spPr>\n'
            f'          <c:marker><c:symbol val="none"/></c:marker>\n'
            f'          <c:xVal>'
            f'<c:numRef><c:f>{_ref(date_col)}</c:f>{empty_cache}</c:numRef>'
            f'</c:xVal>\n'
            f'          <c:yVal>'
            f'<c:numRef><c:f>{_ref(col_ltr)}</c:f>{empty_cache}</c:numRef>'
            f'</c:yVal>\n'
            f'          <c:smooth val="1"/>\n'
            f'        </c:ser>'
        )

    # Build the 20 flow series (primary axis) and 20 pressure series (secondary axis)
    flow_xml = "\n".join(
        _ser(n, FLOW_COLORS[n], get_column_letter(COL_FLOW_ADJ_BASE + n))
        for n in range(MAX_FLOW)
    )
    pres_xml = "\n".join(
        _ser(MAX_FLOW + n, PRES_COLORS[n], get_column_letter(COL_PRES_ADJ_BASE + n))
        for n in range(MAX_PRES)
    )

    return f"""\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"
              xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
              xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <c:lang val="en-US"/>
  <c:style val="10"/>
  <c:chart>
    <c:title>
      <c:tx><c:rich><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Flow &amp; Pressure Analysis</a:t></a:r></a:p></c:rich></c:tx>
      <c:overlay val="0"/>
    </c:title>
    <c:autoTitleDeleted val="0"/>
    <c:plotArea>
      <c:layout/>
      <c:scatterChart>
        <c:scatterStyle val="smoothMarker"/>
        <c:varyColors val="0"/>
{flow_xml}
        <c:axId val="1001"/>
        <c:axId val="1002"/>
      </c:scatterChart>
      <c:scatterChart>
        <c:scatterStyle val="smoothMarker"/>
        <c:varyColors val="0"/>
{pres_xml}
        <c:axId val="1001"/>
        <c:axId val="1003"/>
      </c:scatterChart>
      <c:valAx>
        <c:axId val="1001"/>
        <c:scaling><c:orientation val="minMax"/></c:scaling>
        <c:delete val="0"/>
        <c:axPos val="b"/>
        <c:numFmt formatCode="d/m/yy h:mm" sourceLinked="0"/>
        <c:majorTickMark val="out"/>
        <c:minorTickMark val="none"/>
        <c:tickLblPos val="nextTo"/>
        <c:crossAx val="1002"/>
        <c:crosses val="autoZero"/>
        <c:crossBetween val="midCat"/>
      </c:valAx>
      <c:valAx>
        <c:axId val="1002"/>
        <c:scaling><c:orientation val="minMax"/></c:scaling>
        <c:delete val="0"/>
        <c:axPos val="l"/>
        <c:majorGridlines>
          <c:spPr><a:ln w="9525"><a:solidFill><a:srgbClr val="D9D9D9"/></a:solidFill></a:ln></c:spPr>
        </c:majorGridlines>
        <c:title>
          <c:tx><c:rich><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Flow Adjusted</a:t></a:r></a:p></c:rich></c:tx>
          <c:overlay val="0"/>
        </c:title>
        <c:numFmt formatCode="General" sourceLinked="1"/>
        <c:majorTickMark val="out"/>
        <c:minorTickMark val="none"/>
        <c:tickLblPos val="nextTo"/>
        <c:crossAx val="1001"/>
        <c:crosses val="autoZero"/>
        <c:crossBetween val="midCat"/>
      </c:valAx>
      <c:valAx>
        <c:axId val="1003"/>
        <c:scaling><c:orientation val="minMax"/></c:scaling>
        <c:delete val="0"/>
        <c:axPos val="r"/>
        <c:majorGridlines>
          <c:spPr><a:ln w="9525"><a:solidFill><a:srgbClr val="F4B183"/></a:solidFill><a:prstDash val="dash"/></a:ln></c:spPr>
        </c:majorGridlines>
        <c:title>
          <c:tx><c:rich><a:bodyPr/><a:lstStyle/><a:p><a:r><a:t>Pressure Adjusted</a:t></a:r></a:p></c:rich></c:tx>
          <c:overlay val="0"/>
        </c:title>
        <c:numFmt formatCode="General" sourceLinked="1"/>
        <c:majorTickMark val="out"/>
        <c:minorTickMark val="none"/>
        <c:tickLblPos val="nextTo"/>
        <c:crossAx val="1001"/>
        <c:crosses val="max"/>
        <c:crossBetween val="midCat"/>
      </c:valAx>
    </c:plotArea>
    <c:plotVisOnly val="1"/>
    <c:dispBlanksAs val="gap"/>
    <c:showDLblsOverMax val="0"/>
  </c:chart>
</c:chartSpace>"""


def _build_correct_drawing_xml():
    """Return valid spreadsheet-drawing XML anchoring chart1 at CHART_ANCHOR.

    Uses twoCellAnchor (the format Excel itself always produces for charts) with
    editAs="oneCell" so the chart moves with its top-left cell but keeps a fixed
    size.  Three additional details that real Excel drawings always include —
    and that cause Excel to discard the drawing when absent:

      1.  <a:graphicFrameLocks noGrp="1"/> inside cNvGraphicFramePr
      2.  Explicit xmlns declarations directly on the <c:chart> element
      3.  twoCellAnchor instead of oneCellAnchor
    """
    from openpyxl.utils import column_index_from_string

    # ── from position ──────────────────────────────────────────────────────────
    from_col = column_index_from_string(
        CHART_ANCHOR.rstrip("0123456789")) - 1              # 0-based
    from_row = int("".join(c for c in CHART_ANCHOR if c.isdigit())) - 1  # 0-based

    # ── to position (approximate, based on layout row heights + col widths) ────
    # Column widths set in build_dashboard:  G-U ≈ 13 chars wide
    #   pixels = INT(MDW * chars + 5)  with MDW=7  → 96 px for 13 chars
    #   EMU    = pixels * 9525         → 915 600 EMU per 13-char column
    col_emu  = 96 * 9525                # ≈ 915 600 EMU per data column
    cx = int(CHART_WIDTH_CM  * 360_000)  # 7 200 000 EMU
    cy = int(CHART_HEIGHT_CM * 360_000)  # 5 040 000 EMU

    # Accumulate column widths from from_col until we exceed cx
    to_col    = from_col
    remaining = cx
    while remaining > col_emu:
        remaining -= col_emu
        to_col += 1
    to_col_off = remaining  # EMU offset within the last column

    # Row heights used in build_dashboard (in points → EMU = pt * 12700):
    #   Row 1  (TITLE_ROW)      : 30 pt
    #   Rows 2-22 (SEL_HDR–SEL_END): 22 pt (hdr) + 24 pt each × 20 rows
    #   Row 23 (NOTE_ROW)       : 44 pt
    #   Row 24 (DATA_BANNER)    : 20 pt
    #   Row 25 (DATA_HDR_ROW)   : 22 pt
    #   Row 26+ (DATA_START)    : 15 pt
    row_heights_pt = (
        [30]                                       # row 1
        + [22]                                     # row 2 (SEL_HDR_ROW)
        + [24] * max(MAX_FLOW, MAX_PRES)           # rows 3-22 (SEL selector rows)
        + [44]                                     # NOTE_ROW
        + [20]                                     # DATA_BANNER_ROW
        + [22]                                     # DATA_HDR_ROW
        + [15] * 300                               # data rows, generous upper bound
    )
    to_row    = from_row
    remaining = cy
    idx       = from_row
    while remaining > 0 and idx < len(row_heights_pt):
        h = row_heights_pt[idx] * 12700
        if remaining <= h:
            break
        remaining -= h
        to_row = idx + 1
        idx   += 1
    to_row_off = remaining

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<xdr:wsDr'
        ' xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        ' xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">\n'
        '  <xdr:twoCellAnchor editAs="oneCell">\n'
        f'    <xdr:from>'
        f'<xdr:col>{from_col}</xdr:col><xdr:colOff>0</xdr:colOff>'
        f'<xdr:row>{from_row}</xdr:row><xdr:rowOff>0</xdr:rowOff>'
        f'</xdr:from>\n'
        f'    <xdr:to>'
        f'<xdr:col>{to_col}</xdr:col><xdr:colOff>{to_col_off}</xdr:colOff>'
        f'<xdr:row>{to_row}</xdr:row><xdr:rowOff>{to_row_off}</xdr:rowOff>'
        f'</xdr:to>\n'
        '    <xdr:graphicFrame macro="">\n'
        '      <xdr:nvGraphicFramePr>\n'
        '        <xdr:cNvPr id="2" name="Chart 1"/>\n'
        # graphicFrameLocks is required by Excel; an empty cNvGraphicFramePr
        # causes the drawing to be silently removed on open.
        '        <xdr:cNvGraphicFramePr>'
        '<a:graphicFrameLocks noGrp="1"/>'
        '</xdr:cNvGraphicFramePr>\n'
        '      </xdr:nvGraphicFramePr>\n'
        # xfrm children must be present; position/size come from from+to above.
        '      <xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm>\n'
        '      <a:graphic>\n'
        '        <a:graphicData'
        ' uri="http://schemas.openxmlformats.org/drawingml/2006/chart">\n'
        # Explicit namespace declarations on c:chart are required by Excel's
        # strict validator; inheriting them from the root is insufficient.
        '          <c:chart'
        ' xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
        ' r:id="rId1"/>\n'
        '        </a:graphicData>\n'
        '      </a:graphic>\n'
        '    </xdr:graphicFrame>\n'
        '    <xdr:clientData/>\n'
        '  </xdr:twoCellAnchor>\n'
        '</xdr:wsDr>'
    )


def _build_correct_drawing_rels():
    """Return a drawing relationship file using a relative (not absolute) target path."""
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
        '  <Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"'
        ' Target="../charts/chart1.xml"/>\n'
        '</Relationships>'
    )


def _patch_chart_xml(xlsx_path):
    """Replace chart1.xml, drawing1.xml, and drawing1.xml.rels in the xlsx.

    Two separate OOXML issues must both be fixed for Excel to accept the chart:

    Issue 1 — chart1.xml: CT_LineSer element order
      The OOXML schema requires <c:smooth> to appear AFTER <c:cat> and <c:val>
      inside <c:ser>.  If it appears before them Excel's strict validator rejects
      chart1.xml and cascade-removes drawing1.xml (which references the rejected
      chart part).  _build_correct_chart_xml() places <c:smooth> correctly.

    Issue 2 — drawing1.xml: missing required macro attribute
      openpyxl's native drawing omits the required macro="" attribute on
      <graphicFrame>.  The OOXML CT_GraphicalObjectFrame schema declares this
      attribute as use="required".  Without it Excel discards the drawing.
      _build_correct_drawing_xml() generates a conformant twoCellAnchor with
      macro="", <a:graphicFrameLocks noGrp="1"/>, and proper xfrm children.
    """
    correct_chart   = _build_correct_chart_xml().encode("utf-8")
    correct_drawing = _build_correct_drawing_xml().encode("utf-8")
    correct_drel    = _build_correct_drawing_rels().encode("utf-8")

    tmp_path = xlsx_path + ".patching"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            # OPC spec requires [Content_Types].xml to be the first entry
            ct_item = None
            for item in zin.infolist():
                if item.filename == "[Content_Types].xml":
                    ct_item = item
                    break
            if ct_item:
                zout.writestr(ct_item, zin.read(ct_item.filename))

            for item in zin.infolist():
                if item.filename == "[Content_Types].xml":
                    continue  # already written first
                if item.filename == "xl/charts/chart1.xml":
                    zout.writestr(item, correct_chart)
                elif item.filename == "xl/drawings/drawing1.xml":
                    zout.writestr(item, correct_drawing)
                elif item.filename == "xl/drawings/_rels/drawing1.xml.rels":
                    zout.writestr(item, correct_drel)
                else:
                    zout.writestr(item, zin.read(item.filename))
    os.replace(tmp_path, xlsx_path)


# ── MOD output sheets ──────────────────────────────────────────────────────────

def build_mod_sheet(ws, title):
    """Create an empty wide-format MOD sheet matching the raw-data layout.

    Row 1:   "Date" in A1  (SaveToMOD adds sensor-code headers to B1, C1, …)
    Row 2+:  date/value data written by the SaveToMOD VBA macro.

    The sheet is intentionally left almost empty so the user can paste data from
    the raw sheets side-by-side. SaveToMOD clears and rewrites it on every run.
    """
    ws.title = title
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 21

    style_header(ws.cell(1, 1), "Date", bg=DARK_BLUE)

    ws.row_dimensions[2].height = 18
    ic = ws.cell(2, 1,
                 value=("Run the SaveToMOD macro to populate this sheet.  "
                        "Format: Date in column A; sensor codes as column headers in row 1."))
    ic.font = Font(italic=True, color=DARK_GRAY, size=9)
    ic.alignment = Alignment(wrap_text=True)


# ── Instructions sheet ─────────────────────────────────────────────────────────

def build_instructions(ws):
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 115
    ws.sheet_view.showGridLines = False

    def section(row, text, bg=DARK_BLUE, sz=11, height=26):
        ws.row_dimensions[row].height = height
        c = ws.cell(row, 1, value=text)
        c.font = Font(bold=True, color=WHITE, size=sz)
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def body(row, text, fg="000000", sz=10, indent=0, bold=False,
             italic=False, bg=None, height=17):
        ws.row_dimensions[row].height = height
        c = ws.cell(row, 1, value=text)
        c.font = Font(color=fg, size=sz, bold=bold, italic=italic)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=indent, wrap_text=True)
        if bg:
            c.fill = PatternFill(fill_type="solid", fgColor=bg)

    def blank(row, h=8):
        ws.row_dimensions[row].height = h

    r = 1

    section(r, "Flow & Pressure Dashboard — Instructions", bg=DARK_BLUE, sz=14, height=32)
    r += 1; blank(r); r += 1

    # ── 1. Quick start ────────────────────────────────────────────────────────
    section(r, "1.  QUICK START  (works immediately — no setup needed)", bg=MID_BLUE)
    r += 1
    for line in [
        "Step 1:  Paste your flow data into 'Raw Flow Data' (delete sample rows, keep Row 1 headers).",
        "Step 2:  Paste your pressure data into 'Raw Pressure Data' (same format).",
        "Step 3:  Go to the Dashboard sheet.",
        "Step 4:  Rows 3-22 (col B) are the 20 flow selectors — pick a sensor from the dropdown.",
        "         Leave unused rows blank to hide that series.",
        "Step 5:  Rows 3-22 (col G) are the 20 pressure selectors — same idea.",
        "Step 6:  Adjust the Scale (col C) for each flow row and the Offset (col H) for",
        "         each pressure row independently.  Default Scale = 1.000, Offset = 0.000.",
        "Step 7:  The input cell for each series is coloured to match its chart line.",
        "         Flow lines and pressure lines are both solid.",
        "Step 8:  Use the Chart Controls panel (cols K-L, top right of the Dashboard):",
        "         • Start Date / End Date  — enter dates to filter the formula table and chart.",
        "           Leave blank to show all available data.  Dates must exist in 'Raw Flow Data'.",
        "Step 9:  Each flow row (col D) and each pressure row (col I) has its own Δt cell.",
        "         Enter an integer to shift that series in time:",
        "         +2 = read from 2 timesteps later;  -3 = read from 3 timesteps earlier.",
        "         Use this to align sensors with different transit / delay times.",
        "Step 10: To save a single sensor click its 💾 cell (col E for flow, col J for pressure).",
        "         After setting up the VBA (see section 4) a single click writes only that",
        "         sensor's adjusted data into the MOD sheet without touching other columns.",
        "Step 11: Click '💾 Save Rest' (top-right, row 7) to save all remaining sensors",
        "         that have not yet been individually saved to the MOD sheets.",
        "Step 12: To rebuild both MOD sheets completely from scratch, run SaveToMOD.",
        "         SaveToMOD reads the COMPLETE raw dataset, applies Scale / Offset / Δt,",
        "         and clears/rewrites MOD Flow and MOD Pressure entirely.",
        "",
        "NOTE:  Up to 20 flow series (left Y-axis, blue/teal shades — solid lines) and 20",
        "       pressure series (right Y-axis, warm/cool shades — solid lines) shown simultaneously.",
        "",
        "NOTE:  After pasting your own data, right-click each Name cell → Data Validation",
        "       → update the Source to cover your column range, e.g. 'Raw Flow Data'!$B$1:$BZ$1",
        "",
        f"NOTE:  The formula table covers {DATA_ROWS} rows. For longer datasets, select",
        f"       that range and copy-paste downward as far as needed.",
        "",
        "NOTE:  -999 values are treated as no-data and are excluded from all calculations.",
    ]:
        body(r, line, indent=2)
        r += 1
    blank(r); r += 1

    # ── 2. Power Query (optional enhancement) ─────────────────────────────────
    section(r, "2.  POWER QUERY SETUP  (optional — recommended for very large datasets)",
            bg=DARK_ORANGE)
    r += 1
    for line in [
        "The Raw data sheets are already set up as Excel Tables (FlowData, PressureData).",
        "Power Query can load these tables, unpivot them, and merge them for use in PivotTables.",
        "",
        "Step 1:  Data tab → Get Data → From Table/Range → select the FlowData table.",
        "Step 2:  In Power Query Editor: select the Date column, then Home → Unpivot Other Columns.",
        "         Rename 'Attribute' → 'Flow Name',  'Value' → 'Flow Value'.",
        "Step 3:  Close & Load To… → Only Create Connection.  Name the query  FlowLong.",
        "Step 4:  Repeat for PressureData.  Name the query  PressureLong.",
        "Step 5:  Merge the two queries on Date + Name to get a combined table.",
        "Step 6:  Add calculated columns:  Flow Adjusted = [Flow Value] × scaling_factor",
        "                                  Pressure Adjusted = [Pressure Value] + offset",
        "Step 7:  Load the merged query to a sheet and build a PivotTable + Slicer on top of it.",
        "",
        "After pasting new data:  Data tab → Refresh All  (Ctrl+Alt+F5).",
    ]:
        body(r, line, indent=2)
        r += 1
    blank(r); r += 1

    # ── 3. PivotTable & Slicer ────────────────────────────────────────────────
    section(r, "3.  PIVOTTABLE + SLICER  (optional — for interactive multi-flow comparison)",
            bg=GREEN_DARK)
    r += 1
    for line in [
        "Once the Power Query merged table is loaded to a sheet:",
        "  • Insert → PivotTable",
        "  • Rows: Date    Values: Flow Adjusted, Pressure Adjusted",
        "  • PivotTable Analyze → Insert Slicer → tick 'Flow Name' → OK",
        "  • Click a flow name in the Slicer to filter instantly",
        "  • Insert → PivotChart → Line → add Secondary Axis to the Pressure series",
    ]:
        body(r, line, indent=2)
        r += 1
    blank(r); r += 1

    # ── 4. VBA Save macros ────────────────────────────────────────────────────
    section(r, "4.  VBA SAVE MACROS — two steps:", bg=PURPLE)
    r += 1
    for line in [
        "Step A:  Press Alt+F11 to open the VBA editor.",
        "Step B:  Click Insert → Module and paste the code block labelled",
        "         'STANDARD MODULE CODE' below.  This adds SaveToMOD,",
        "         SaveOneSensorToMOD, and SaveRemainingToMOD.",
        "Step C:  In the Project tree (left panel) double-click the Dashboard",
        "         sheet module (usually 'Sheet1 (Dashboard)') and paste the",
        "         code block labelled 'DASHBOARD SHEET MODULE CODE' below.",
        "         This makes the 💾 cells and the 'Save Rest' cell clickable.",
        "Step D:  Close the VBA editor and save the file as .xlsm.",
    ]:
        body(r, line, indent=2)
        r += 1
    blank(r); r += 1

    vba = r"""' ===========================================================================
' STANDARD MODULE CODE - paste into a new Module (Alt+F11 -> Insert -> Module)
' ===========================================================================

' SaveToMOD - clears and rewrites BOTH MOD sheets from scratch using the
'             current Scale, Offset, and Dt settings for ALL sensors.
'
' Dashboard selector rows 3-22:
'   Col B (2) = Flow sensor name    Col C (3) = Flow scale multiplier
'   Col D (4) = Flow Dt (row offset)
'   Col G (7) = Pressure sensor name  Col H (8) = Pressure offset addend
'   Col I (9) = Pressure Dt (row offset)
' ===========================================================================
Sub SaveToMOD()

    Const SEL_START      As Long = 3    ' first selector row on Dashboard
    Const MAX_ROWS       As Long = 20   ' up to 20 flow / 20 pressure series
    Const FLOW_SEL_COL   As Long = 2    ' B - flow sensor name
    Const FLOW_SCALE_COL As Long = 3    ' C - scale multiplier
    Const FLOW_DT_COL    As Long = 4    ' D - timestep offset
    Const PRES_SEL_COL   As Long = 7    ' G - pressure sensor name
    Const PRES_OFF_COL   As Long = 8    ' H - offset addend
    Const PRES_DT_COL    As Long = 9    ' I - timestep offset

    Dim wsDash    As Worksheet
    Dim wsRawFlow As Worksheet
    Dim wsRawPres As Worksheet
    Dim wsModFlow As Worksheet
    Dim wsModPres As Worksheet

    On Error Resume Next
    Set wsDash    = Worksheets("Dashboard")
    Set wsRawFlow = Worksheets("Raw Flow Data")
    Set wsRawPres = Worksheets("Raw Pressure Data")
    Set wsModFlow = Worksheets("MOD Flow")
    Set wsModPres = Worksheets("MOD Pressure")
    On Error GoTo 0

    If wsDash Is Nothing Or wsRawFlow Is Nothing Or wsRawPres Is Nothing Or _
       wsModFlow Is Nothing Or wsModPres Is Nothing Then
        MsgBox "Could not find required sheets (Dashboard, Raw Flow Data, " & _
               "Raw Pressure Data, MOD Flow, MOD Pressure).", vbCritical, "Sheet Missing"
        Exit Sub
    End If

    Dim i As Long, j As Long, k As Long, col As Long, srcRow As Long
    Dim flowName(0 To 19)  As String
    Dim flowScale(0 To 19) As Double
    Dim flowDt(0 To 19)    As Long
    Dim flowCol(0 To 19)   As Long   ' column in Raw Flow Data (0 = not found)
    Dim presName(0 To 19)  As String
    Dim presOff(0 To 19)   As Double
    Dim presDt(0 To 19)    As Long
    Dim presCol(0 To 19)   As Long   ' column in Raw Pressure Data (0 = not found)

    ' -- Read selector settings from Dashboard ----------------------------------
    For i = 0 To MAX_ROWS - 1
        flowName(i)  = Trim(wsDash.Cells(SEL_START + i, FLOW_SEL_COL).Value)
        flowScale(i) = CDbl(wsDash.Cells(SEL_START + i, FLOW_SCALE_COL).Value)
        If flowScale(i) = 0 Then flowScale(i) = 1
        flowDt(i)    = CLng(wsDash.Cells(SEL_START + i, FLOW_DT_COL).Value)
        presName(i)  = Trim(wsDash.Cells(SEL_START + i, PRES_SEL_COL).Value)
        presOff(i)   = CDbl(wsDash.Cells(SEL_START + i, PRES_OFF_COL).Value)
        presDt(i)    = CLng(wsDash.Cells(SEL_START + i, PRES_DT_COL).Value)
    Next i

    ' -- Locate sensor columns in raw sheet headers (row 1) --------------------
    Dim lastFlowHdrCol As Long
    Dim lastPresHdrCol As Long
    lastFlowHdrCol = wsRawFlow.Cells(1, wsRawFlow.Columns.Count).End(xlToLeft).Column
    lastPresHdrCol = wsRawPres.Cells(1, wsRawPres.Columns.Count).End(xlToLeft).Column

    For i = 0 To MAX_ROWS - 1
        flowCol(i) = 0
        presCol(i) = 0
        If flowName(i) <> "" Then
            For k = 2 To lastFlowHdrCol
                If Trim(wsRawFlow.Cells(1, k).Value) = flowName(i) Then
                    flowCol(i) = k
                    Exit For
                End If
            Next k
        End If
        If presName(i) <> "" Then
            For k = 2 To lastPresHdrCol
                If Trim(wsRawPres.Cells(1, k).Value) = presName(i) Then
                    presCol(i) = k
                    Exit For
                End If
            Next k
        End If
    Next i

    Dim lastFlowRow As Long
    Dim lastPresRow As Long
    lastFlowRow = wsRawFlow.Cells(wsRawFlow.Rows.Count, 1).End(xlUp).Row
    lastPresRow = wsRawPres.Cells(wsRawPres.Rows.Count, 1).End(xlUp).Row

    Application.ScreenUpdating = False
    Dim totalSaved As Long: totalSaved = 0
    Dim rawVal As Variant
    Dim dtVal  As Variant
    Dim outRow As Long

    ' -- MOD Flow: process every row in Raw Flow Data ---------------------------
    wsModFlow.Cells.ClearContents
    wsModFlow.Cells(1, 1).Value = "Date"
    col = 2
    For i = 0 To MAX_ROWS - 1
        If flowName(i) <> "" Then
            wsModFlow.Cells(1, col).Value = flowName(i)
            col = col + 1
        End If
    Next i

    outRow = 2
    For j = 2 To lastFlowRow
        dtVal = wsRawFlow.Cells(j, 1).Value
        If IsEmpty(dtVal) Or dtVal = "" Then GoTo NextFlowRow
        wsModFlow.Cells(outRow, 1).Value        = dtVal
        wsModFlow.Cells(outRow, 1).NumberFormat = "DD/MM/YYYY HH:MM"
        col = 2
        For i = 0 To MAX_ROWS - 1
            If flowName(i) <> "" Then
                If flowCol(i) > 0 Then
                    srcRow = j - flowDt(i)
                    If srcRow >= 2 And srcRow <= lastFlowRow Then
                        rawVal = wsRawFlow.Cells(srcRow, flowCol(i)).Value
                        If IsNumeric(rawVal) And CDbl(rawVal) <> -999 Then
                            wsModFlow.Cells(outRow, col).Value        = CDbl(rawVal) * flowScale(i)
                            wsModFlow.Cells(outRow, col).NumberFormat = "0.000"
                            totalSaved = totalSaved + 1
                        End If
                    End If
                End If
                col = col + 1
            End If
        Next i
        outRow = outRow + 1
NextFlowRow:
    Next j

    ' -- MOD Pressure: process every row in Raw Pressure Data ------------------
    wsModPres.Cells.ClearContents
    wsModPres.Cells(1, 1).Value = "Date"
    col = 2
    For i = 0 To MAX_ROWS - 1
        If presName(i) <> "" Then
            wsModPres.Cells(1, col).Value = presName(i)
            col = col + 1
        End If
    Next i

    outRow = 2
    For j = 2 To lastPresRow
        dtVal = wsRawPres.Cells(j, 1).Value
        If IsEmpty(dtVal) Or dtVal = "" Then GoTo NextPresRow
        wsModPres.Cells(outRow, 1).Value        = dtVal
        wsModPres.Cells(outRow, 1).NumberFormat = "DD/MM/YYYY HH:MM"
        col = 2
        For i = 0 To MAX_ROWS - 1
            If presName(i) <> "" Then
                If presCol(i) > 0 Then
                    srcRow = j - presDt(i)
                    If srcRow >= 2 And srcRow <= lastPresRow Then
                        rawVal = wsRawPres.Cells(srcRow, presCol(i)).Value
                        If IsNumeric(rawVal) And CDbl(rawVal) <> -999 Then
                            wsModPres.Cells(outRow, col).Value        = CDbl(rawVal) + presOff(i)
                            wsModPres.Cells(outRow, col).NumberFormat = "0.000"
                            totalSaved = totalSaved + 1
                        End If
                    End If
                End If
                col = col + 1
            End If
        Next i
        outRow = outRow + 1
NextPresRow:
    Next j

    Application.ScreenUpdating = True

    If totalSaved = 0 Then
        MsgBox "No data values were saved - check that sensor names match the " & _
               "column headers in Raw Flow Data / Raw Pressure Data.", _
               vbExclamation, "Nothing Saved"
    Else
        MsgBox "Saved " & totalSaved & " values covering the complete dataset " & _
               "(all dates, not just the current dashboard window).", _
               vbInformation, "Save Complete"
    End If

End Sub

' ===========================================================================
' SaveOneSensorToMOD - saves adjusted data for ONE sensor row to the
'                      appropriate MOD sheet WITHOUT clearing other columns.
'
' isFlow  : True = Flow sensor, False = Pressure sensor
' sRow    : selector row index 1-20
' Silent  : suppress the confirmation MsgBox (used by SaveRemainingToMOD)
' ===========================================================================
Sub SaveOneSensorToMOD(isFlow As Boolean, sRow As Long, _
                       Optional Silent As Boolean = False)

    Const SEL_START    As Long = 3
    Const FLOW_NAME    As Long = 2   ' B
    Const FLOW_SCALE   As Long = 3   ' C
    Const FLOW_DT      As Long = 4   ' D
    Const PRES_NAME    As Long = 7   ' G
    Const PRES_OFF     As Long = 8   ' H
    Const PRES_DT      As Long = 9   ' I

    Dim wsDash  As Worksheet
    Dim wsRaw   As Worksheet
    Dim wsMod   As Worksheet

    Set wsDash = Worksheets("Dashboard")

    Dim sensorName As String
    Dim scale      As Double
    Dim offset     As Double
    Dim dt         As Long
    Dim dashRow    As Long
    dashRow = SEL_START + sRow - 1

    If isFlow Then
        sensorName = Trim(wsDash.Cells(dashRow, FLOW_NAME).Value)
        If sensorName = "" Then
            If Not Silent Then
                MsgBox "No sensor selected in Flow row " & sRow, vbExclamation
            End If
            Exit Sub
        End If
        scale  = CDbl(wsDash.Cells(dashRow, FLOW_SCALE).Value)
        If scale = 0 Then scale = 1
        dt     = CLng(wsDash.Cells(dashRow, FLOW_DT).Value)
        Set wsRaw = Worksheets("Raw Flow Data")
        Set wsMod = Worksheets("MOD Flow")
    Else
        sensorName = Trim(wsDash.Cells(dashRow, PRES_NAME).Value)
        If sensorName = "" Then
            If Not Silent Then
                MsgBox "No sensor selected in Pressure row " & sRow, vbExclamation
            End If
            Exit Sub
        End If
        offset = CDbl(wsDash.Cells(dashRow, PRES_OFF).Value)
        dt     = CLng(wsDash.Cells(dashRow, PRES_DT).Value)
        Set wsRaw = Worksheets("Raw Pressure Data")
        Set wsMod = Worksheets("MOD Pressure")
    End If

    ' --- Find sensor column in raw sheet ---
    Dim lastRawHdrCol As Long
    lastRawHdrCol = wsRaw.Cells(1, wsRaw.Columns.Count).End(xlToLeft).Column
    Dim rawSensorCol As Long: rawSensorCol = 0
    Dim k As Long
    For k = 2 To lastRawHdrCol
        If Trim(wsRaw.Cells(1, k).Value) = sensorName Then
            rawSensorCol = k
            Exit For
        End If
    Next k
    If rawSensorCol = 0 Then
        If Not Silent Then
            MsgBox "Sensor '" & sensorName & "' not found in " & wsRaw.Name, _
                   vbExclamation, "Sensor Not Found"
        End If
        Exit Sub
    End If

    ' --- Ensure MOD sheet has a Date header ---
    If wsMod.Cells(1, 1).Value = "" Then wsMod.Cells(1, 1).Value = "Date"

    ' --- Extend date column in MOD to match raw sheet ---
    Dim lastRawRow As Long
    lastRawRow = wsRaw.Cells(wsRaw.Rows.Count, 1).End(xlUp).Row
    Dim lastModRow As Long
    lastModRow = wsMod.Cells(wsMod.Rows.Count, 1).End(xlUp).Row

    Dim j As Long
    Dim dtVal As Variant
    Dim startFill As Long
    startFill = IIf(lastModRow < 2, 2, lastModRow + 1)
    If startFill <= lastRawRow Then
        For j = startFill To lastRawRow
            dtVal = wsRaw.Cells(j, 1).Value
            If Not (IsEmpty(dtVal) Or dtVal = "") Then
                wsMod.Cells(j, 1).Value        = dtVal
                wsMod.Cells(j, 1).NumberFormat = "DD/MM/YYYY HH:MM"
            End If
        Next j
    End If

    ' --- Find or create column for this sensor in MOD ---
    Dim lastModHdrCol As Long
    lastModHdrCol = wsMod.Cells(1, wsMod.Columns.Count).End(xlToLeft).Column
    Dim modSensorCol As Long: modSensorCol = 0
    For k = 2 To lastModHdrCol
        If Trim(wsMod.Cells(1, k).Value) = sensorName Then
            modSensorCol = k
            Exit For
        End If
    Next k
    If modSensorCol = 0 Then
        modSensorCol = lastModHdrCol + 1
        wsMod.Cells(1, modSensorCol).Value = sensorName
    End If

    ' --- Write adjusted sensor data ---
    Application.ScreenUpdating = False
    Dim totalSaved As Long: totalSaved = 0
    Dim rawVal As Variant
    Dim srcRow As Long

    For j = 2 To lastRawRow
        dtVal = wsRaw.Cells(j, 1).Value
        If IsEmpty(dtVal) Or dtVal = "" Then GoTo NextSensorRow
        srcRow = j - dt
        If srcRow >= 2 And srcRow <= lastRawRow Then
            rawVal = wsRaw.Cells(srcRow, rawSensorCol).Value
            If IsNumeric(rawVal) And CDbl(rawVal) <> -999 Then
                If isFlow Then
                    wsMod.Cells(j, modSensorCol).Value = CDbl(rawVal) * scale
                Else
                    wsMod.Cells(j, modSensorCol).Value = CDbl(rawVal) + offset
                End If
                wsMod.Cells(j, modSensorCol).NumberFormat = "0.000"
                totalSaved = totalSaved + 1
            End If
        End If
NextSensorRow:
    Next j

    Application.ScreenUpdating = True

    If Not Silent Then
        MsgBox "Saved " & totalSaved & " values for '" & sensorName & "'.", _
               vbInformation, "Sensor Saved"
    End If
End Sub

' ===========================================================================
' SaveRemainingToMOD - saves all sensors selected on the Dashboard that do
'                      NOT yet have a column in the corresponding MOD sheet.
'                      Run this after finishing individual per-row saves.
' ===========================================================================
Sub SaveRemainingToMOD()

    Const SEL_START  As Long = 3
    Const MAX_ROWS   As Long = 20
    Const FLOW_NAME  As Long = 2    ' B
    Const PRES_NAME  As Long = 7    ' G

    Dim wsDash    As Worksheet
    Dim wsModFlow As Worksheet
    Dim wsModPres As Worksheet

    Set wsDash    = Worksheets("Dashboard")
    Set wsModFlow = Worksheets("MOD Flow")
    Set wsModPres = Worksheets("MOD Pressure")

    Dim i As Long, k As Long
    Dim sensorName As String
    Dim found As Boolean
    Dim lastHdrCol As Long
    Dim savedCount As Long: savedCount = 0

    For i = 0 To MAX_ROWS - 1
        ' --- Flow ---
        sensorName = Trim(wsDash.Cells(SEL_START + i, FLOW_NAME).Value)
        If sensorName <> "" Then
            found = False
            lastHdrCol = wsModFlow.Cells(1, wsModFlow.Columns.Count).End(xlToLeft).Column
            For k = 2 To lastHdrCol
                If Trim(wsModFlow.Cells(1, k).Value) = sensorName Then
                    found = True
                    Exit For
                End If
            Next k
            If Not found Then
                SaveOneSensorToMOD True, i + 1, True
                savedCount = savedCount + 1
            End If
        End If

        ' --- Pressure ---
        sensorName = Trim(wsDash.Cells(SEL_START + i, PRES_NAME).Value)
        If sensorName <> "" Then
            found = False
            lastHdrCol = wsModPres.Cells(1, wsModPres.Columns.Count).End(xlToLeft).Column
            For k = 2 To lastHdrCol
                If Trim(wsModPres.Cells(1, k).Value) = sensorName Then
                    found = True
                    Exit For
                End If
            Next k
            If Not found Then
                SaveOneSensorToMOD False, i + 1, True
                savedCount = savedCount + 1
            End If
        End If
    Next i

    If savedCount = 0 Then
        MsgBox "All sensors already saved - nothing to do.", _
               vbInformation, "Save Rest"
    Else
        MsgBox "Saved remaining " & savedCount & " sensor(s) to MOD sheets.", _
               vbInformation, "Save Rest Complete"
    End If
End Sub"""

    body(r, vba, sz=9, bg=LIGHT_GRAY, height=17)
    r += 1
    blank(r); r += 1

    # ── Dashboard sheet module code ───────────────────────────────────────────
    section(r, "DASHBOARD SHEET MODULE CODE  (paste into the Dashboard sheet's code module)",
            bg=PURPLE)
    r += 1
    body(r,
         "In the VBA editor Project tree, double-click 'Sheet1 (Dashboard)' and paste the "
         "code below.  This makes the 💾 cells (cols E and J, rows 3-22) and the "
         "'💾 Save Rest' cell (col K row 7) respond to single clicks.",
         indent=2, italic=True, fg=DARK_GRAY)
    r += 1
    blank(r); r += 1

    vba_sheet = r"""' ===========================================================================
' DASHBOARD SHEET MODULE CODE
' Paste into the Dashboard sheet module (double-click Sheet1 in Project tree)
' ===========================================================================
Private Sub Worksheet_SelectionChange(ByVal Target As Range)

    Const FLOW_SAVE_COL    As Long = 5    ' E - flow [S] cells
    Const PRES_SAVE_COL    As Long = 10   ' J - pres [S] cells
    Const SAVE_REST_COL    As Long = 11   ' K - Save Rest button (merged K:L)
    Const SEL_START        As Long = 3    ' first selector row
    Const SEL_END          As Long = 22   ' last selector row
    Const SAVE_REST_ROW    As Long = 7    ' Save Rest row

    If Target.Count > 1 Then Exit Sub

    If Target.Column = FLOW_SAVE_COL And _
       Target.Row >= SEL_START And Target.Row <= SEL_END Then
        Application.EnableEvents = False
        Target.Offset(0, -1).Select   ' move focus so next click re-fires
        Application.EnableEvents = True
        SaveOneSensorToMOD True, Target.Row - SEL_START + 1

    ElseIf Target.Column = PRES_SAVE_COL And _
           Target.Row >= SEL_START And Target.Row <= SEL_END Then
        Application.EnableEvents = False
        Target.Offset(0, -1).Select
        Application.EnableEvents = True
        SaveOneSensorToMOD False, Target.Row - SEL_START + 1

    ElseIf Target.Column = SAVE_REST_COL And Target.Row = SAVE_REST_ROW Then
        Application.EnableEvents = False
        Target.Offset(1, 0).Select
        Application.EnableEvents = True
        SaveRemainingToMOD
    End If
End Sub"""

    body(r, vba_sheet, sz=9, bg=LIGHT_GRAY, height=17)
    r += 1
    blank(r); r += 1

    # ── 5. Data format reference ──────────────────────────────────────────────
    section(r, "5.  DATA FORMAT REFERENCE", bg=DARK_GRAY)
    r += 1
    for line in [
        "Both Raw Data sheets AND MOD sheets use this wide format:",
        "",
        "    Date               | AL012       | AL013       | AL014       | ...  ",
        "    01/12/2026 00:00   | 3.168205    | 2.204250    | 2.665153    | ...  ",
        "    01/12/2026 00:15   | 3.190769    | 2.225250    | 2.681334    | ...  ",
        "",
        "  Dates are displayed in UK format: DD/MM/YYYY HH:MM",
        "",
        "• Column A must contain a proper Date/Time value (not text).",
        "• Flow and pressure column names can be any mix of letters and numbers.",
        "• The names in Raw Flow Data and Raw Pressure Data do NOT need to match",
        "  — you select each independently on the Dashboard.",
        "• Use -999 for missing/no-data values — they are excluded from all calculations.",
        "• Data can be any length: months of 15-minute data = thousands of rows.",
        "• To paste your own data into a raw sheet: delete the sample data rows",
        "  (keep row 1 headers), then paste starting from row 2.",
        "",
        "Dashboard data table (rows 26+):",
        "  Col A     = Date",
        "  Cols B-U  = Flow 1-20 Adjusted  (Name in B3-B22 × Scale in C3-C22 + Δt in D3-D22)",
        "  Cols V-AO = Pres 1-20 Adjusted  (Name in G3-G22 + Offset in H3-H22 + Δt in I3-I22)",
        "",
        "MOD Flow / MOD Pressure:",
        "  SaveToMOD       — clears and rewrites both sheets from scratch.",
        "  💾 per-row cells — add/update just that sensor's column (non-destructive).",
        "  💾 Save Rest    — saves all sensors not yet present in the MOD sheets.",
        "  Row 1  = Date | SensorCode1 | SensorCode2 | ...  (headers)",
        "  Row 2+ = adjusted values in matching columns",
        "",
        "Leave a Name cell blank to hide that series (formula returns empty, not plotted).",
    ]:
        body(r, line, indent=2)
        r += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    ws_flow  = wb.active
    ws_pres  = wb.create_sheet()
    ws_dash  = wb.create_sheet()
    ws_mod_f = wb.create_sheet()
    ws_mod_p = wb.create_sheet()
    ws_instr = wb.create_sheet()

    build_raw_sheet(ws_flow, "Raw Flow Data",     "FlowData",     FLOW_ROWS, SAMPLE_DATES)
    build_raw_sheet(ws_pres, "Raw Pressure Data", "PressureData", PRES_ROWS, SAMPLE_DATES)
    build_dashboard(ws_dash, FLOW_NAMES)
    build_mod_sheet(ws_mod_f, "MOD Flow")
    build_mod_sheet(ws_mod_p, "MOD Pressure")
    build_instructions(ws_instr)

    wb.calculation.calcMode    = "auto"
    wb.calculation.fullCalcOnLoad = True

    out = "Flow_Pressure_Dashboard.xlsx"
    wb.save(out)
    _patch_chart_xml(out)
    print(f"Generated: {out}")


if __name__ == "__main__":
    main()
